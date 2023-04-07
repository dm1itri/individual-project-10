from data import db_session
from data.game import Game
from data.player import Player
from data.history_move import HistoryMove
from data.question import Question

from openpyxl import load_workbook


def clear_table(session):
    session.query(Game).delete()
    session.query(Player).delete()
    session.query(HistoryMove).delete()
    session.query(Question).delete()
    session.commit()


def add_game(session, number_of_players, max_number_of_questions):
    game = Game(current_player=0, number_of_players=number_of_players, max_number_of_questions=max_number_of_questions)
    session.add(game)
    session.commit()
    return game


def add_null_history_move(session, game_id):
    null_history_move = HistoryMove(game_id=game_id, number_history=0, number_move=-1, number_steps=0)
    session.add(null_history_move)
    session.commit()


def add_players(session, game, number_of_players):
    for i in range(number_of_players):
        player = Player(current_position=0, skipping_move=False, number_move=i, game=game)
        session.add(player)
    session.commit()


def clear_game_players(session, game_id):
    game = session.get(Game, game_id)
    game.current_player = 0
    players = session.query(Player).filter(Player.game_id == game.id).all()
    for player in players:
        player.current_position = 0
        player.skipping_move = False
        player.is_occupied = False
    session.query(HistoryMove).filter(HistoryMove.game_id == game_id).delete()
    history = HistoryMove(game_id=game_id)
    session.add(history)
    session.commit()


def add_question(session, question_d):
    question = Question()
    question.type_question = question_d['type_question']
    question.question = question_d['question']
    question.answer_correct = question_d['answer_correct']
    question.answer_2 = question_d['answer_2']
    question.answer_3 = question_d['answer_3']
    question.answer_4 = question_d['answer_4']
    session.add(question)
    session.commit()


def read_queastions_xlsx(filename):
    wb = load_workbook(filename)
    keys_d = ['question', 'answer_correct', 'answer_2', 'answer_3', 'answer_4']
    all_questions = []
    for name in ['Биология', 'История', 'География']:
        sheet = wb[name]
        for i in range(1, 6):
            d = {'type_question': name}
            for j in range(1, 6):
                d[keys_d[j - 1]] = sheet.cell(row=i, column=j).value
            all_questions.append(d)
    with db_session.create_session() as session:
        for question in all_questions:
            add_question(session, question)